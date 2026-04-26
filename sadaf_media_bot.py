#!/usr/bin/env python3
"""
Sadaf Media Video Studio — Telegram Buyurtma Boti
Ishlatish: pip install pyTelegramBotAPI openpyxl
Keyin: python sadaf_media_bot.py
"""

import telebot
import json
import os
import io
from telebot import types
from datetime import datetime

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

BOT_TOKEN = "8213889849:AAHqpH_BGU0iaWns8YMW1j5hg7wbNJfd0Ao"
MANAGER_CHAT_ID = 7948989650
ORDERS_FILE = "orders.json"
USERS_FILE  = "users.json"
PROMO_FILE  = "promo_codes.json"

bot = telebot.TeleBot(BOT_TOKEN)
user_data = {}

XIZMATLAR        = ["🎬 Toy videosi", "📢 Reklama video", "📸 Foto sessiya", "🏢 Korporativ video"]
TOY_TURLARI      = ["💍 Nikoh", "🍽 Banket", "✂️ Xatna", "👶 Chaqaloq", "🕌 Haj/Umra", "🎂 Tug'ilgan kun"]
TOY_QOSHIMCHA_LIST = ["📷 Fotograf", "🏗 Kran", "📅 Yana bir kun", "🎥 +1 Kamera"]
QOSHIMCHA_LIST   = ["📷 Fotograf", "🏗 Kran"]

TOY_NARX = {
    "💍 Nikoh": 700_000, "🍽 Banket": 700_000, "✂️ Xatna": 700_000,
    "👶 Chaqaloq": 500_000, "🕌 Haj/Umra": 500_000, "🎂 Tug'ilgan kun": 500_000,
}
XIZMAT_NARX = {
    "📸 Foto sessiya": 500_000,
    "📢 Reklama video": 700_000,
    "🏢 Korporativ video": 500_000,
}
TOY_QOSHIMCHA_NARX = {
    "📷 Fotograf": 200_000, "🏗 Kran": 1_000_000,
    "📅 Yana bir kun": 700_000, "🎥 +1 Kamera": 700_000,
}
QOSHIMCHA_NARX = {"📷 Fotograf": 200_000, "🏗 Kran": 1_000_000}


def fmt(n):
    return f"{n:,}".replace(",", " ") + " so'm"


# ══════════════════════════════════════════════
#  PROMO KODLAR
# ══════════════════════════════════════════════

def load_promos():
    if not os.path.exists(PROMO_FILE):
        return {}
    with open(PROMO_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_promos(data):
    with open(PROMO_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_promo(code):
    promos = load_promos()
    return promos.get(code.upper().strip())

def use_promo(code):
    promos = load_promos()
    key = code.upper().strip()
    if key in promos and promos[key].get("active", True):
        if promos[key].get("bir_marta"):
            promos[key]["active"] = False
            save_promos(promos)
        return promos[key]
    return None


# ══════════════════════════════════════════════
#  USERS JSON
# ══════════════════════════════════════════════

def load_users():
    if not os.path.exists(USERS_FILE):
        return {}
    with open(USERS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_users(users):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

def register_user(cid, telefon, user_info):
    users = load_users()
    existing = users.get(str(cid), {})
    users[str(cid)] = {
        "user_id": cid,
        "user_name": user_info,
        "telefon": telefon,
        "sana": existing.get("sana", datetime.now().strftime("%d.%m.%Y %H:%M")),
        "blocked": existing.get("blocked", False),
    }
    save_users(users)

def get_user(cid):
    return load_users().get(str(cid))

def is_blocked(cid):
    u = get_user(cid)
    return bool(u and u.get("blocked"))

def block_user(cid):
    users = load_users()
    if str(cid) in users:
        users[str(cid)]["blocked"] = True
        save_users(users)

def unblock_user(cid):
    users = load_users()
    if str(cid) in users:
        users[str(cid)]["blocked"] = False
        save_users(users)

def restore_user_data(cid):
    saved = get_user(cid)
    if saved and cid not in user_data:
        user_data[cid] = {
            "step": "done",
            "telefon": saved.get("telefon", ""),
            "qoshimcha": [],
        }
    return saved


# ══════════════════════════════════════════════
#  ORDERS JSON
# ══════════════════════════════════════════════

def load_orders():
    if not os.path.exists(ORDERS_FILE):
        return []
    with open(ORDERS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_orders(orders):
    with open(ORDERS_FILE, "w", encoding="utf-8") as f:
        json.dump(orders, f, ensure_ascii=False, indent=2)

def get_order_by_id(oid):
    for o in load_orders():
        if o["id"] == oid:
            return o
    return None

def cancel_order(oid):
    orders = load_orders()
    for o in orders:
        if o["id"] == oid:
            o["status"] = "bekor"
    save_orders(orders)

def delete_order(oid):
    orders = load_orders()
    save_orders([o for o in orders if o["id"] != oid])

def update_order_narx(oid, yangi_narx):
    orders = load_orders()
    for o in orders:
        if o["id"] == oid:
            o["jami"] = yangi_narx
            old_blok = o.get("narx_blok", "")
            if "💵" in old_blok:
                o["narx_blok"] = old_blok.split("💵")[0] + f"💵 *Jami: {fmt(yangi_narx)}* _(admin o'zgartirdi)_\n"
            else:
                o["narx_blok"] = f"\n💵 *Jami: {fmt(yangi_narx)}* _(admin o'zgartirdi)_\n"
    save_orders(orders)

def get_user_orders(user_id):
    return [o for o in load_orders() if o.get("user_id") == user_id]

def get_all_users_from_orders():
    seen = {}
    for o in load_orders():
        uid = o.get("user_id")
        if uid and uid not in seen:
            seen[uid] = o.get("user_name", str(uid))
    return seen


# ══════════════════════════════════════════════
#  STATISTIKA
# ══════════════════════════════════════════════

def get_stats_text():
    orders = load_orders()
    users  = load_users()

    jami_buyurtma = len(orders)
    aktiv         = sum(1 for o in orders if o.get("status") == "aktiv")
    bekor         = sum(1 for o in orders if o.get("status") == "bekor")
    jami_daromad  = sum(o.get("jami", 0) for o in orders if o.get("status") == "aktiv")
    jami_users    = len(users)
    blocked_count = sum(1 for u in users.values() if u.get("blocked"))

    xizmat_count = {}
    for o in orders:
        x = o.get("xizmat_str", "Noma'lum").split("(")[0].strip()
        xizmat_count[x] = xizmat_count.get(x, 0) + 1
    xizmat_lines = "\n".join(
        [f"  • {k}: {v} ta" for k, v in sorted(xizmat_count.items(), key=lambda i: -i[1])]
    )

    bugun = datetime.now().strftime("%d.%m.%Y")
    bugun_count = sum(1 for o in orders if o.get("vaqt", "").startswith(bugun))

    return (
        "📊 *Statistika*\n\n"
        f"📦 Jami buyurtmalar: *{jami_buyurtma} ta*\n"
        f"  🟢 Aktiv: *{aktiv} ta*\n"
        f"  🔴 Bekor: *{bekor} ta*\n"
        f"📅 Bugungi buyurtmalar: *{bugun_count} ta*\n\n"
        f"💵 Jami daromad (aktiv): *{fmt(jami_daromad)}*\n\n"
        f"👥 Jami foydalanuvchilar: *{jami_users} ta*\n"
        f"  🚫 Bloklangan: *{blocked_count} ta*\n\n"
        f"🎬 Xizmat bo'yicha:\n{xizmat_lines or '  —'}"
    )


# ══════════════════════════════════════════════
#  EXCEL / CSV EKSPORT
# ══════════════════════════════════════════════

def export_orders_excel():
    orders = load_orders()
    if not orders:
        return None, "xlsx"

    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Buyurtmalar"
        headers = ["#", "Holat", "Xizmat", "Sana", "Manzil", "Telefon", "Mijoz", "Jami (so'm)", "Vaqt"]
        ws.append(headers)
        for o in orders:
            telefon_val = o.get("telefon", "")
            ws.append([
                o.get("id"), o.get("status"), o.get("xizmat_str"),
                o.get("sana"), o.get("joy_text"), telefon_val,
                o.get("user_name"), o.get("jami", 0), o.get("vaqt"),
            ])
            # Telefon ustunini (F) matn sifatida belgilash
            last_row = ws.max_row
            cell = ws.cell(row=last_row, column=6)
            cell.value = telefon_val
            cell.number_format = '@'
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf, "xlsx"
    else:
        lines = ["#,Holat,Xizmat,Sana,Manzil,Telefon,Mijoz,Jami,Vaqt"]
        for o in orders:
            row = [
                str(o.get("id","")), o.get("status",""), o.get("xizmat_str",""),
                o.get("sana",""), o.get("joy_text",""), o.get("telefon",""),
                o.get("user_name",""), str(o.get("jami",0)), o.get("vaqt",""),
            ]
            lines.append(",".join(f'"{c}"' for c in row))
        buf = io.BytesIO("\n".join(lines).encode("utf-8-sig"))
        buf.seek(0)
        return buf, "csv"


# ══════════════════════════════════════════════
#  ADMIN CHECK
# ══════════════════════════════════════════════

def is_admin(cid):
    return cid == MANAGER_CHAT_ID


# ══════════════════════════════════════════════
#  ADMIN PANEL
# ══════════════════════════════════════════════

def show_admin_panel(cid):
    user_data[cid] = {"step": "admin"}
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add(
        types.KeyboardButton("📋 Buyurtmalar"),
        types.KeyboardButton("👥 Foydalanuvchilar"),
        types.KeyboardButton("📊 Statistika"),
        types.KeyboardButton("📤 Excel eksport"),
        types.KeyboardButton("📢 Broadcast"),
        types.KeyboardButton("💬 Chat"),
        types.KeyboardButton("🎁 Promo kodlar"),
    )
    bot.send_message(cid, "👨‍💼 *Admin panel*\n\nNimani ko'rmoqchisiz?",
                     parse_mode="Markdown", reply_markup=markup)


def show_orders_list(cid):
    orders = load_orders()
    if not orders:
        bot.send_message(cid, "📭 Hozircha buyurtmalar yo'q.")
        return
    user_data[cid]["step"] = "admin_orders"
    markup = types.InlineKeyboardMarkup()
    for o in reversed(orders):
        icon  = "🔴" if o.get("status") == "bekor" else "🟢"
        label = f"{icon} #{o['id']} | {o.get('xizmat_str','—')} | {o.get('sana','—')} | {o.get('telefon','—')}"
        markup.add(types.InlineKeyboardButton(label, callback_data=f"order_{o['id']}"))
    bot.send_message(cid, "📋 *Barcha buyurtmalar:*", parse_mode="Markdown", reply_markup=markup)


def show_order_detail(cid, oid):
    o = get_order_by_id(oid)
    if not o:
        bot.send_message(cid, "Buyurtma topilmadi.")
        return
    status = "🔴 Bekor qilingan" if o.get("status") == "bekor" else "🟢 Aktiv"
    text = (
        f"📋 *Buyurtma #{o['id']}*\n\n"
        f"🎬 Xizmat: {o.get('xizmat_str','—')}\n"
        f"📅 Sana: {o.get('sana','—')}\n"
        f"📍 Manzil: {o.get('joy_text','—')}\n"
        f"➕ Qo'shimcha:\n{o.get('qoshimcha_str', "Yo'q")}\n"
        f"📱 Telefon: {o.get('telefon','—')}\n"
        f"{o.get('narx_blok','')}\n"
        f"👤 Mijoz: {o.get('user_name','—')}\n"
        f"🕐 Vaqt: {o.get('vaqt','—')}\n"
        f"📊 Holat: {status}"
    )
    markup = types.InlineKeyboardMarkup()
    if o.get("status") != "bekor":
        markup.add(types.InlineKeyboardButton("❌ Bekor qilish", callback_data=f"cancel_{oid}"))
    markup.add(types.InlineKeyboardButton("✏️ Narxni o'zgartirish", callback_data=f"edit_narx_{oid}"))
    markup.add(types.InlineKeyboardButton("🗑 O'chirish", callback_data=f"admin_delete_{oid}"))
    markup.add(types.InlineKeyboardButton("⬅️ Ortga", callback_data="back_orders"))
    if o.get("joy_lat") and o.get("joy_lon"):
        bot.send_location(cid, o["joy_lat"], o["joy_lon"])
    bot.send_message(cid, text, parse_mode="Markdown", reply_markup=markup)


def show_users_list(cid):
    users = load_users()
    if not users:
        bot.send_message(cid, "👥 Hozircha foydalanuvchilar yo'q.")
        return
    markup = types.InlineKeyboardMarkup()
    for uid, u in users.items():
        blocked = u.get("blocked", False)
        icon = "🚫" if blocked else "✅"
        name = u.get("user_name", uid)[:30]
        markup.add(types.InlineKeyboardButton(
            f"{icon} {name}", callback_data=f"admin_user_{uid}"
        ))
    bot.send_message(cid, "👥 *Foydalanuvchilar ro'yxati:*\n✅ Aktiv  🚫 Bloklangan",
                     parse_mode="Markdown", reply_markup=markup)


def show_user_detail_admin(cid, target_uid):
    users = load_users()
    u = users.get(str(target_uid))
    if not u:
        bot.send_message(cid, "Foydalanuvchi topilmadi.")
        return
    blocked = u.get("blocked", False)
    orders  = get_user_orders(int(target_uid))
    jami_s  = sum(o.get("jami", 0) for o in orders if o.get("status") == "aktiv")
    text = (
        f"👤 *Foydalanuvchi:*\n\n"
        f"🏷 Ism: {u.get('user_name','—')}\n"
        f"📱 Telefon: {u.get('telefon','—')}\n"
        f"🕐 Ro'yxatdan: {u.get('sana','—')}\n"
        f"📦 Buyurtmalar: {len(orders)} ta\n"
        f"💵 Jami to'lov: {fmt(jami_s)}\n"
        f"📊 Holat: {'🚫 Bloklangan' if blocked else '✅ Aktiv'}"
    )
    markup = types.InlineKeyboardMarkup()
    if blocked:
        markup.add(types.InlineKeyboardButton("✅ Blokdan chiqarish", callback_data=f"unblock_{target_uid}"))
    else:
        markup.add(types.InlineKeyboardButton("🚫 Bloklash", callback_data=f"block_{target_uid}"))
    markup.add(types.InlineKeyboardButton("💬 Xabar yuborish", callback_data=f"chat_user_{target_uid}"))
    markup.add(types.InlineKeyboardButton("⬅️ Ortga", callback_data="back_users"))
    bot.send_message(cid, text, parse_mode="Markdown", reply_markup=markup)


def start_broadcast(cid):
    user_data[cid]["step"] = "admin_broadcast"
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton("⬅️ Ortga"))
    bot.send_message(cid,
        "📢 *Broadcast*\n\nBarcha foydalanuvchilarga yubormoqchi bo'lgan xabarni yozing:\n_(Matn, rasm yoki video)_",
        parse_mode="Markdown", reply_markup=markup)


def do_broadcast(cid, message):
    users = load_users()
    ok, fail = 0, 0
    for uid_str, u in users.items():
        if u.get("blocked"):
            continue
        try:
            uid = int(uid_str)
            if message.content_type == "text":
                bot.send_message(uid, message.text)
            elif message.content_type == "photo":
                bot.send_photo(uid, message.photo[-1].file_id, caption=message.caption or "")
            elif message.content_type == "video":
                bot.send_video(uid, message.video.file_id, caption=message.caption or "")
            ok += 1
        except Exception:
            fail += 1
    bot.send_message(cid, f"✅ Broadcast tugadi!\n\n📨 Yuborildi: {ok} ta\n❌ Xato: {fail} ta")
    show_admin_panel(cid)


# ══════════════════════════════════════════════
#  CALLBACK HANDLER
# ══════════════════════════════════════════════

@bot.callback_query_handler(func=lambda c: True)
def handle_callback(call):
    cid  = call.message.chat.id
    data = call.data

    if data.startswith("order_"):
        oid = int(data.split("_")[1])
        show_order_detail(cid, oid)
        bot.answer_callback_query(call.id)

    elif data.startswith("cancel_"):
        oid = int(data.split("_")[1])
        o   = get_order_by_id(oid)
        cancel_order(oid)
        bot.answer_callback_query(call.id, "Bekor qilindi")
        if o:
            try:
                bot.send_message(o["user_id"],
                    "Kechirasiz, buyurtmangiz *bekor qilindi*.\n"
                    "Qo'shimcha ma'lumot: +998974787478\nAdmin @doniyorbekgulomov0",
                    parse_mode="Markdown")
            except Exception:
                pass
        bot.edit_message_text(f"Buyurtma #{oid} bekor qilindi.", cid, call.message.message_id)

    elif data == "back_orders":
        bot.answer_callback_query(call.id)
        show_orders_list(cid)

    elif data.startswith("edit_narx_"):
        oid = int(data.split("_")[2])
        user_data[cid]["step"]       = "admin_edit_narx"
        user_data[cid]["edit_order"] = oid
        bot.answer_callback_query(call.id)
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton("Ortga"))
        bot.send_message(cid,
            f"✏️ *#{oid}-buyurtma* uchun yangi narxni kiriting:\n_(Faqat raqam, masalan: 1500000)_",
            parse_mode="Markdown", reply_markup=markup)

    elif data.startswith("admin_user_"):
        uid = data.split("_")[2]
        bot.answer_callback_query(call.id)
        show_user_detail_admin(cid, uid)

    elif data == "back_users":
        bot.answer_callback_query(call.id)
        show_users_list(cid)

    elif data.startswith("block_"):
        target = int(data.split("_")[1])
        block_user(target)
        bot.answer_callback_query(call.id, "Bloklandi")
        try:
            bot.send_message(target,
                "🚫 *Siz bloklandi*\n\n"
                "Botdan foydalana olmaysiz.\n"
                "Qo'shimcha ma'lumot uchun:\n"
                "Admin @doniyorbekgulomov0\nTel: +998974787478",
                parse_mode="Markdown")
        except Exception:
            pass
        try:
            bot.edit_message_text(
                f"Foydalanuvchi {target} bloklandi.", cid, call.message.message_id)
        except Exception:
            pass
        show_users_list(cid)

    elif data.startswith("unblock_"):
        target = int(data.split("_")[1])
        unblock_user(target)
        bot.answer_callback_query(call.id, "Blokdan chiqarildi")
        try:
            bot.send_message(target,
                "✅ *Blokingiz olib tashlandi!*\n\n"
                "Endi botdan foydalanishingiz mumkin. /start bosing.",
                parse_mode="Markdown")
        except Exception:
            pass
        try:
            bot.edit_message_text(
                f"Foydalanuvchi {target} blokdan chiqarildi.", cid, call.message.message_id)
        except Exception:
            pass
        show_users_list(cid)

    elif data.startswith("chat_user_"):
        uid   = int(data.split("_")[2])
        users = load_users()
        uname = users.get(str(uid), {}).get("user_name", str(uid))
        user_data[cid] = {"step": "admin_chat_send", "chat_target": uid, "chat_target_name": uname}
        bot.answer_callback_query(call.id)
        bot.send_message(cid,
            f"💬 *{uname}* ga xabar yuboring:\n_(Matn, rasm yoki video)_",
            parse_mode="Markdown",
            reply_markup=types.ReplyKeyboardMarkup(resize_keyboard=True).add(
                types.KeyboardButton("⬅️ Ortga")))

    elif data.startswith("admin_delete_"):
        oid = int(data.split("_")[2])
        bot.answer_callback_query(call.id)
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("✅ Ha, o'chirish", callback_data=f"admin_confirm_delete_{oid}"),
            types.InlineKeyboardButton("❌ Bekor", callback_data=f"order_{oid}")
        )
        bot.send_message(cid,
            f"⚠️ *#{oid}-buyurtmani* o'chirishni tasdiqlaysizmi?",
            parse_mode="Markdown", reply_markup=markup)

    elif data.startswith("admin_confirm_delete_"):
        oid = int(data.split("_")[3])
        o   = get_order_by_id(oid)
        delete_order(oid)
        bot.answer_callback_query(call.id, "O'chirildi ✅")
        bot.edit_message_text(
            f"🗑 *#{oid}-buyurtma o'chirildi.*", cid, call.message.message_id,
            parse_mode="Markdown")
        show_orders_list(cid)

    elif data.startswith("my_order_"):
        oid = int(data.split("_")[2])
        bot.answer_callback_query(call.id)
        show_user_order_detail(cid, oid)

    elif data.startswith("delete_my_"):
        oid = int(data.split("_")[2])
        bot.answer_callback_query(call.id)
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("✅ Ha, o'chirish", callback_data=f"confirm_delete_{oid}"),
            types.InlineKeyboardButton("❌ Yo'q", callback_data=f"my_order_{oid}")
        )
        bot.send_message(cid,
            f"⚠️ *#{oid}-buyurtmani* o'chirishni tasdiqlaysizmi?",
            parse_mode="Markdown", reply_markup=markup)

    elif data.startswith("confirm_delete_"):
        oid = int(data.split("_")[2])
        o   = get_order_by_id(oid)
        if not o or o.get("user_id") != cid:
            bot.answer_callback_query(call.id, "Ruxsat yo'q!")
            return
        delete_order(oid)
        bot.answer_callback_query(call.id, "O'chirildi")
        bot.edit_message_text(
            f"#{oid}-buyurtma o'chirildi.", cid, call.message.message_id)
        try:
            bot.send_message(MANAGER_CHAT_ID,
                f"🗑 *#{oid}-buyurtma mijoz tomonidan o'chirildi.*\n"
                f"👤 {o.get('user_name','—')} | 📱 {o.get('telefon','—')}",
                parse_mode="Markdown")
        except Exception:
            pass
        show_user_orders(cid)

    elif data == "back_my_orders":
        bot.answer_callback_query(call.id)
        show_user_orders(cid)


# ══════════════════════════════════════════════
#  /start
# ══════════════════════════════════════════════

@bot.message_handler(commands=["start", "buyurtma"])
def start(message):
    cid = message.chat.id
    if is_admin(cid):
        show_admin_panel(cid)
        return
    if is_blocked(cid):
        bot.send_message(cid,
            "🚫 *Siz bloklandi*\n\n"
            "Botdan foydalana olmaysiz.\n"
            "Qo'shimcha ma'lumot uchun:\n"
            "Admin @doniyorbekgulomov0\nTel: +998974787478",
            parse_mode="Markdown")
        return
    restore_user_data(cid)
    ask_telefon_first(cid)


# ══════════════════════════════════════════════
#  KONTAKT
# ══════════════════════════════════════════════

@bot.message_handler(content_types=["contact"])
def handle_contact(message):
    cid = message.chat.id
    if is_blocked(cid):
        return
    if cid not in user_data or user_data[cid].get("step") != "telefon":
        return
    phone = message.contact.phone_number
    if not phone.startswith("+"):
        phone = "+" + phone
    user_data[cid]["telefon"] = phone
    ask_xizmat(cid)


# ══════════════════════════════════════════════
#  LOKATSIYA
# ══════════════════════════════════════════════

@bot.message_handler(content_types=["location"])
def handle_location(message):
    cid = message.chat.id
    if is_blocked(cid):
        return
    if cid not in user_data or user_data[cid].get("step") != "joy":
        return
    user_data[cid]["joy_lat"] = message.location.latitude
    user_data[cid]["joy_lon"] = message.location.longitude
    user_data[cid]["step"]    = "joy_text"
    bot.send_message(cid,
        "✅ Lokatsiya qabul qilindi!\n\n📝 Endi manzilni *matn* ko'rinishida ham yozing:",
        parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())


# ══════════════════════════════════════════════
#  MEDIA
# ══════════════════════════════════════════════

@bot.message_handler(content_types=["photo", "video", "document"])
def handle_media(message):
    cid   = message.chat.id
    state = user_data.get(cid, {})

    if is_admin(cid) and state.get("step") == "admin_broadcast":
        do_broadcast(cid, message)
        return

    if is_admin(cid) and state.get("step") == "admin_chat_send":
        target = state.get("chat_target")
        if not target:
            return
        try:
            if message.photo:
                bot.send_photo(target, message.photo[-1].file_id, caption=message.caption or "")
            elif message.video:
                bot.send_video(target, message.video.file_id, caption=message.caption or "")
            elif message.document:
                bot.send_document(target, message.document.file_id, caption=message.caption or "")
            bot.send_message(cid, "✅ Yuborildi!")
        except Exception as e:
            bot.send_message(cid, f"❌ Xato: {e}")
        show_admin_panel(cid)


# ══════════════════════════════════════════════
#  ASOSIY XABAR HANDLER
# ══════════════════════════════════════════════

@bot.message_handler(func=lambda m: True)
def handle_message(message):
    cid  = message.chat.id
    text = message.text.strip() if message.text else ""

    if is_admin(cid):
        handle_admin_message(cid, text, message)
        return

    if is_blocked(cid):
        bot.send_message(cid,
            "🚫 *Siz bloklandi.*\n"
            "Admin @doniyorbekgulomov0 | Tel: +998974787478",
            parse_mode="Markdown")
        return

    if text == "🔄 Yangi buyurtma":
        ask_telefon_first(cid)
        return

    if text == "📋 Buyurtmalarim":
        show_user_orders(cid)
        return

    if cid not in user_data:
        restore_user_data(cid)

    if cid not in user_data:
        ask_telefon_first(cid)
        return

    state = user_data[cid]
    step  = state.get("step")

    if step == "telefon":
        state["telefon"] = text
        ask_xizmat(cid)

    elif step == "xizmat":
        if text not in XIZMATLAR:
            bot.send_message(cid, "Iltimos, quyidagi tugmalardan birini tanlang ⬇️")
            return
        state["xizmat"] = text
        if text == "🎬 Toy videosi":
            state["step"] = "toy_turi"
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
            markup.add(*[types.KeyboardButton(x) for x in TOY_TURLARI])
            bot.send_message(cid, f"✅ *{text}* tanlandi!\n\nQanday toy turi?",
                             parse_mode="Markdown", reply_markup=markup)
        elif text == "📸 Foto sessiya":
            state["skip_qoshimcha"] = True
            state["step"] = "sana"
            bot.send_message(cid,
                f"✅ *{text}* tanlandi! — {fmt(XIZMAT_NARX[text])}\n\n📅 Tadbir sanasini kiriting:",
                parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
        else:
            state["step"] = "sana"
            narx_text = f" — {fmt(XIZMAT_NARX[text])}" if text in XIZMAT_NARX else ""
            bot.send_message(cid,
                f"✅ *{text}* tanlandi!{narx_text}\n\n📅 Tadbir sanasini kiriting:\n_(Masalan: 5-may-2026)_",
                parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())

    elif step == "toy_turi":
        if text not in TOY_TURLARI:
            bot.send_message(cid, "Iltimos, tugmalardan birini tanlang ⬇️")
            return
        state["toy_turi"] = text
        state["step"]     = "sana"
        bot.send_message(cid,
            f"✅ *{text}* tanlandi! — {fmt(TOY_NARX.get(text, 0))}\n\n📅 Tadbir sanasini kiriting:",
            parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())

    elif step == "sana":
        state["sana"] = text
        state["step"] = "joy"
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        markup.add(types.KeyboardButton("📍 Lokatsiya yuborish", request_location=True))
        bot.send_message(cid,
            f"📅 Sana: *{text}*\n\n📍 Tadbir qayerda bo'ladi?\nLokatsiya yuboring:",
            parse_mode="Markdown", reply_markup=markup)

    elif step == "joy_text":
        state["joy_text"] = text
        if state.get("skip_qoshimcha"):
            ask_promo(cid, state)
        else:
            ask_qoshimcha(cid, state)

    elif step == "qoshimcha_confirm":
        if text == "✅ Shu yetarli":
            ask_promo(cid, state)
        elif text == "➕ Yana xizmat qo'shish":
            ask_qoshimcha(cid, state)
        else:
            bot.send_message(cid, "Iltimos, tugmalardan birini tanlang ⬇️")

    elif step == "qoshimcha":
        is_toy    = state.get("xizmat") == "🎬 Toy videosi"
        qlist     = TOY_QOSHIMCHA_LIST if is_toy else QOSHIMCHA_LIST
        available = [x for x in qlist if x not in state["qoshimcha"]]
        if text not in available:
            bot.send_message(cid, "Iltimos, tugmalardan birini tanlang ⬇️")
            return
        state["qoshimcha"].append(text)
        narx_map = TOY_QOSHIMCHA_NARX if is_toy else QOSHIMCHA_NARX
        narx     = narx_map.get(text, 0)
        state["step"] = "qoshimcha_confirm"
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        markup.add(
            types.KeyboardButton("➕ Yana xizmat qo'shish"),
            types.KeyboardButton("✅ Shu yetarli")
        )
        bot.send_message(cid,
            f"✅ *{text}* qo'shildi! — {fmt(narx)}\n\nYana qo'shimcha xizmat kerakmi?",
            parse_mode="Markdown", reply_markup=markup)

    elif step == "promo":
        if text == "⏭ O'tkazib yuborish":
            state.pop("promo_kod", None)
            state.pop("chegirma_foiz", None)
            finalize(cid, state)
        else:
            promo = get_promo(text)
            if promo and promo.get("active"):
                state["promo_kod"] = text.upper().strip()
                state["chegirma_foiz"] = promo["foiz"]
                bot.send_message(cid,
                    f"🎉 *{text.upper()}* kodi qabul qilindi!\n💸 Chegirma: *{promo['foiz']}%*\n\nBuyurtma rasmiylashtirilmoqda...",
                    parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
                finalize(cid, state)
            else:
                bot.send_message(cid,
                    "❌ Bu promo kod mavjud emas yoki muddati tugagan.\n\nQaytadan kiriting yoki o'tkazib yuboring:")

    else:
        ask_telefon_first(cid)


# ══════════════════════════════════════════════
#  ADMIN XABAR HANDLER
# ══════════════════════════════════════════════

def handle_admin_message(cid, text, message):
    state = user_data.get(cid, {"step": "admin"})
    step  = state.get("step")

    if step == "admin_broadcast":
        if text == "⬅️ Ortga":
            show_admin_panel(cid)
            return
        do_broadcast(cid, message)
        return

    if step == "admin_chat_send":
        if text == "⬅️ Ortga":
            show_admin_panel(cid)
            return
        target = state.get("chat_target")
        try:
            bot.send_message(target, text)
            bot.send_message(cid, "✅ Xabar yuborildi!")
        except Exception as e:
            bot.send_message(cid, f"❌ Xato: {e}")
        show_admin_panel(cid)
        return

    if step == "admin_promo_create":
        if text in ("⬅️ Ortga", "Ortga"):
            show_promo_panel(cid)
            return
        # Format: KOD:FOIZ  masalan SADAF10:10
        parts = text.upper().strip().split(":")
        if len(parts) != 2 or not parts[1].isdigit():
            bot.send_message(cid,
                "❌ Noto'g'ri format!\n\nTo'g'ri format: *KOD:FOIZ*\nMasalan: `SADAF10:10`",
                parse_mode="Markdown")
            return
        kod, foiz = parts[0], int(parts[1])
        if foiz <= 0 or foiz > 100:
            bot.send_message(cid, "❌ Foiz 1 dan 100 gacha bo'lishi kerak!")
            return
        promos = load_promos()
        promos[kod] = {"foiz": foiz, "active": True, "bir_marta": False}
        save_promos(promos)
        bot.send_message(cid,
            f"✅ *Promo kod yaratildi!*\n\n🎁 Kod: `{kod}`\n💸 Chegirma: *{foiz}%*",
            parse_mode="Markdown")
        show_promo_panel(cid)
        return

    if step == "admin_promo_delete":
        if text in ("⬅️ Ortga", "Ortga"):
            show_promo_panel(cid)
            return
        kod = text.upper().strip()
        promos = load_promos()
        if kod in promos:
            del promos[kod]
            save_promos(promos)
            bot.send_message(cid, f"🗑 *{kod}* kodi o'chirildi.", parse_mode="Markdown")
        else:
            bot.send_message(cid, f"❌ *{kod}* kodi topilmadi.", parse_mode="Markdown")
        show_promo_panel(cid)
        return

    if step == "admin_edit_narx":
        if text in ("⬅️ Ortga", "Ortga"):
            show_admin_panel(cid)
            return
        oid = state.get("edit_order")
        try:
            yangi = int(text.replace(" ", "").replace(",", ""))
            update_order_narx(oid, yangi)
            bot.send_message(cid,
                f"✅ *#{oid}-buyurtma* narxi *{fmt(yangi)}* ga o'zgartirildi!",
                parse_mode="Markdown")
            o = get_order_by_id(oid)
            if o:
                try:
                    bot.send_message(o["user_id"],
                        f"📋 *#{oid}-buyurtmangiz* narxi yangilandi.\n"
                        f"💵 Yangi narx: *{fmt(yangi)}*",
                        parse_mode="Markdown")
                except Exception:
                    pass
        except ValueError:
            bot.send_message(cid, "❌ Faqat raqam kiriting! (masalan: 1500000)")
            return
        show_admin_panel(cid)
        return

    if text == "📋 Buyurtmalar":
        show_orders_list(cid)
    elif text == "👥 Foydalanuvchilar":
        show_users_list(cid)
    elif text == "📊 Statistika":
        bot.send_message(cid, get_stats_text(), parse_mode="Markdown")
    elif text == "📤 Excel eksport":
        buf, ext = export_orders_excel()
        if buf is None:
            bot.send_message(cid, "📭 Buyurtmalar yo'q.")
        else:
            fname = f"buyurtmalar_{datetime.now().strftime('%d_%m_%Y')}.{ext}"
            bot.send_document(cid, (fname, buf),
                caption=f"📊 Buyurtmalar eksporti — {datetime.now().strftime('%d.%m.%Y')}")
    elif text == "📢 Broadcast":
        start_broadcast(cid)
    elif text == "💬 Chat":
        show_chat_users(cid)
    elif text == "🎁 Promo kodlar":
        show_promo_panel(cid)
    elif text in ("➕ Yangi kod yaratish", "🗑 Kodni o'chirish"):
        handle_promo_panel_buttons(cid, text)
    elif text in ("⬅️ Ortga", "/start"):
        show_admin_panel(cid)
    else:
        show_admin_panel(cid)


def show_promo_panel(cid):
    user_data[cid] = {"step": "admin"}
    promos = load_promos()
    if promos:
        lines = "\n".join([
            f"  {'✅' if v.get('active') else '❌'} `{k}` — {v['foiz']}% chegirma"
            for k, v in promos.items()
        ])
        text = f"🎁 *Promo kodlar:*\n\n{lines}"
    else:
        text = "🎁 *Promo kodlar:*\n\nHozircha kodlar yo'q."
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add(
        types.KeyboardButton("➕ Yangi kod yaratish"),
        types.KeyboardButton("🗑 Kodni o'chirish"),
        types.KeyboardButton("⬅️ Ortga"),
    )
    bot.send_message(cid, text, parse_mode="Markdown", reply_markup=markup)


def handle_promo_panel_buttons(cid, text):
    if text == "➕ Yangi kod yaratish":
        user_data[cid] = {"step": "admin_promo_create"}
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton("⬅️ Ortga"))
        bot.send_message(cid,
            "✏️ Yangi promo kodni quyidagi formatda yozing:\n\n"
            "*KOD:FOIZ*\n\nMasalan: `SADAF10:10`\n_(10% chegirma beradi)_",
            parse_mode="Markdown", reply_markup=markup)
    elif text == "🗑 Kodni o'chirish":
        promos = load_promos()
        if not promos:
            bot.send_message(cid, "❌ O'chirish uchun kodlar yo'q.")
            return
        user_data[cid] = {"step": "admin_promo_delete"}
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
        for k in promos:
            markup.add(types.KeyboardButton(k))
        markup.add(types.KeyboardButton("⬅️ Ortga"))
        bot.send_message(cid, "🗑 Qaysi kodni o'chirmoqchisiz?", reply_markup=markup)



    users_dict = get_all_users_from_orders()
    if not users_dict:
        bot.send_message(cid, "👤 Hozircha foydalanuvchilar yo'q.")
        return
    markup = types.InlineKeyboardMarkup()
    for uid, name in users_dict.items():
        markup.add(types.InlineKeyboardButton(f"👤 {name}", callback_data=f"chat_user_{uid}"))
    bot.send_message(cid, "💬 *Kimga xabar yuborasiz?*", parse_mode="Markdown", reply_markup=markup)


# ══════════════════════════════════════════════
#  USER: O'Z BUYURTMALARI
# ══════════════════════════════════════════════

def show_user_orders(cid):
    orders = get_user_orders(cid)
    if not orders:
        bot.send_message(cid, "📭 Sizda hozircha buyurtmalar yo'q.")
        return
    markup = types.InlineKeyboardMarkup()
    for o in reversed(orders):
        icon  = "🔴" if o.get("status") == "bekor" else "🟢"
        label = f"{icon} #{o['id']} | {o.get('xizmat_str','—')} | {o.get('sana','—')}"
        markup.add(types.InlineKeyboardButton(label, callback_data=f"my_order_{o['id']}"))
    bot.send_message(cid, "📋 *Mening buyurtmalarim:*", parse_mode="Markdown", reply_markup=markup)


def show_user_order_detail(cid, oid):
    o = get_order_by_id(oid)
    if not o:
        bot.send_message(cid, "Buyurtma topilmadi.")
        return
    if o.get("user_id") != cid:
        bot.send_message(cid, "⛔ Bu buyurtma sizniki emas.")
        return
    status = "🔴 Bekor qilingan" if o.get("status") == "bekor" else "🟢 Aktiv"
    text = (
        f"📋 *Buyurtma #{o['id']}*\n\n"
        f"🎬 Xizmat: {o.get('xizmat_str','—')}\n"
        f"📅 Sana: {o.get('sana','—')}\n"
        f"📍 Manzil: {o.get('joy_text','—')}\n"
        f"➕ Qo'shimcha:\n{o.get('qoshimcha_str', "Yo'q")}\n"
        f"📱 Telefon: {o.get('telefon','—')}\n"
        f"{o.get('narx_blok','')}\n"
        f"🕐 Vaqt: {o.get('vaqt','—')}\n"
        f"📊 Holat: {status}"
    )
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("🗑 Buyurtmani o'chirish", callback_data=f"delete_my_{oid}"))
    markup.add(types.InlineKeyboardButton("⬅️ Ortga", callback_data="back_my_orders"))
    bot.send_message(cid, text, parse_mode="Markdown", reply_markup=markup)


# ══════════════════════════════════════════════
#  YORDAMCHI FUNKSIYALAR
# ══════════════════════════════════════════════

def ask_telefon_first(cid):
    saved = get_user(cid)
    if saved and saved.get("telefon"):
        user_data[cid] = {"step": "xizmat", "telefon": saved["telefon"], "qoshimcha": []}
        ask_xizmat(cid)
        return
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.add(types.KeyboardButton("📱 Telefon raqamni yuborish", request_contact=True))
    bot.send_message(cid,
        "👋 Salom! *Sadaf Media Video Studio*ga xush kelibsiz!\n\n"
        "Qo'shimcha ma'lumot uchun:\nAdmin @doniyorbekgulomov0\nNomer📞 +998974787478\n\n"
        "📱 Avval telefon raqamingizni yuboring:",
        parse_mode="Markdown", reply_markup=markup)
    user_data[cid] = {"step": "telefon", "qoshimcha": []}


def ask_xizmat(cid):
    user_data[cid]["step"] = "xizmat"
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add(*[types.KeyboardButton(x) for x in XIZMATLAR])
    markup.add(types.KeyboardButton("📋 Buyurtmalarim"))
    bot.send_message(cid, "Qaysi xizmatga buyurtma berishni xohlaysiz?",
                     parse_mode="Markdown", reply_markup=markup)


def ask_qoshimcha(cid, state):
    is_toy    = state.get("xizmat") == "🎬 Toy videosi"
    qlist     = TOY_QOSHIMCHA_LIST if is_toy else QOSHIMCHA_LIST
    narx_map  = TOY_QOSHIMCHA_NARX if is_toy else QOSHIMCHA_NARX
    available = [x for x in qlist if x not in state["qoshimcha"]]
    if not available:
        finalize(cid, state)
        return
    state["step"] = "qoshimcha"
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add(*[types.KeyboardButton(x) for x in available])
    already   = ", ".join(state["qoshimcha"]) if state["qoshimcha"] else "hech biri"
    narx_info = "\n".join([f"• {k}: {fmt(v)}" for k, v in narx_map.items() if k in available])
    bot.send_message(cid,
        f"➕ Qo'shimcha xizmat kerakmi?\n\n{narx_info}\n\n_(Tanlangan: {already})_",
        parse_mode="Markdown", reply_markup=markup)


def ask_promo(cid, state):
    state["step"] = "promo"
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton("⏭ O'tkazib yuborish"))
    bot.send_message(cid,
        "🎁 *Promo kodingiz bormi?*\n\nKodni kiriting yoki o'tkazib yuboring:",
        parse_mode="Markdown", reply_markup=markup)



    toy_turi        = state.get("toy_turi", "")
    xizmat          = state.get("xizmat", "—")
    qoshimcha_list  = state.get("qoshimcha", [])
    is_toy          = xizmat == "🎬 Toy videosi"
    narx_map        = TOY_QOSHIMCHA_NARX if is_toy else QOSHIMCHA_NARX
    asosiy_narx     = TOY_NARX.get(toy_turi, 0) if toy_turi else XIZMAT_NARX.get(xizmat, 0)
    qoshimcha_jami  = sum(narx_map.get(x, 0) for x in qoshimcha_list)
    jami            = asosiy_narx + qoshimcha_jami
    chegirma_foiz   = state.get("chegirma_foiz", 0)
    promo_kod       = state.get("promo_kod", "")
    chegirma_summa  = int(jami * chegirma_foiz / 100) if chegirma_foiz else 0
    jami_chegirma   = jami - chegirma_summa
    xizmat_str      = xizmat + (f" ({toy_turi})" if toy_turi else "")
    q_lines         = "\n".join([f"  • {x}: {fmt(narx_map.get(x,0))}" for x in qoshimcha_list]) or "  Yo'q"

    narx_blok = ""
    if jami > 0:
        narx_blok = f"\n💰 *Narx hisobi:*\n  • Asosiy xizmat: {fmt(asosiy_narx)}\n"
        for x in qoshimcha_list:
            narx_blok += f"  • {x}: {fmt(narx_map.get(x,0))}\n"
        if chegirma_foiz:
            narx_blok += f"  🎁 Promo `{promo_kod}` chegirma ({chegirma_foiz}%): -{fmt(chegirma_summa)}\n"
            narx_blok += f"  ➖➖➖➖➖➖➖\n  💵 *Jami: {fmt(jami_chegirma)}* _(chegirmadan keyin)_\n"
        else:
            narx_blok += f"  ➖➖➖➖➖➖➖\n  💵 *Jami: {fmt(jami)}*\n"

    summary = (
        "📋 *Buyurtma xulosasi:*\n\n"
        f"🎬 Xizmat: {xizmat_str}\n"
        f"📅 Sana: {state.get('sana','—')}\n"
        f"📍 Manzil: {state.get('joy_text','—')}\n"
        f"➕ Qo'shimcha:\n{q_lines}\n"
        f"📱 Telefon: {state.get('telefon','—')}\n"
        f"{narx_blok}"
    )

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add(
        types.KeyboardButton("🔄 Yangi buyurtma"),
        types.KeyboardButton("📋 Buyurtmalarim")
    )
    bot.send_message(cid,
        summary + "\n✅ *Buyurtmangiz qabul qilindi!*\n"
        "Tez orada menejerimiz siz bilan bog'lanadi. 🎉\n"
        "Qo'shimcha ma'lumot uchun:\nAdmin @doniyorbekgulomov0\nNomer📞 +998974787478",
        parse_mode="Markdown", reply_markup=markup)

    try:
        chat      = bot.get_chat(cid)
        username  = f"@{chat.username}" if chat.username else f"ID: {cid}"
        full_name = f"{chat.first_name or ''} {chat.last_name or ''}".strip()
        user_info = f"{full_name} | {username}"
    except Exception:
        user_info = str(cid)

    register_user(cid, state.get("telefon", ""), user_info)

    orders = load_orders()
    new_id = (max([o["id"] for o in orders], default=0)) + 1
    order_obj = {
        "id":            new_id,
        "status":        "aktiv",
        "user_id":       cid,
        "user_name":     user_info,
        "xizmat_str":    xizmat_str,
        "sana":          state.get("sana", "—"),
        "joy_text":      state.get("joy_text", "—"),
        "joy_lat":       state.get("joy_lat"),
        "joy_lon":       state.get("joy_lon"),
        "qoshimcha_str": q_lines,
        "telefon":       state.get("telefon", "—"),
        "narx_blok":     narx_blok,
        "jami":          jami_chegirma if chegirma_foiz else jami,
        "promo_kod":     promo_kod if promo_kod else None,
        "vaqt":          datetime.now().strftime("%d.%m.%Y %H:%M"),
    }
    orders.append(order_obj)
    save_orders(orders)

    if promo_kod:
        use_promo(promo_kod)

    try:
        bot.send_message(MANAGER_CHAT_ID,
            f"🔔 *Yangi buyurtma #{new_id}!*\n\n{summary}\n👤 Mijoz: {user_info}",
            parse_mode="Markdown")
        if state.get("joy_lat") and state.get("joy_lon"):
            bot.send_location(MANAGER_CHAT_ID, state["joy_lat"], state["joy_lon"])
    except Exception as e:
        print(f"Menejer xabari xatosi: {e}")

    user_data[cid] = {"step": "xizmat", "telefon": state.get("telefon", ""), "qoshimcha": []}


print("🚀 Sadaf Media bot ishga tushdi!")
bot.infinity_polling()
